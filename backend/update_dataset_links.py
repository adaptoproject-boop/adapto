import openpyxl
import json
import os

excel_path = r'D:\web\client\sakshi\document\updated dataset.xlsx'
curriculum_path = r'd:\web\client\sakshi\Adaptive-E-Learning-Platform-main (2)\Adaptive-E-Learning-Platform-main\e-learning platform\Adaptive-E-Learning-Platform\backend\data\curriculum.json'
curriculum_new_path = r'd:\web\client\sakshi\Adaptive-E-Learning-Platform-main (2)\Adaptive-E-Learning-Platform-main\e-learning platform\Adaptive-E-Learning-Platform\backend\data\curriculum_new.json'

# New video URLs mapped by the topic names used in curriculum_new.json
new_links_mapping_new = {
    "Hello, Goodbye & Colours": "https://www.youtube.com/watch?v=fN1Cyr0ZK9M",
    "Opposites & Simple Sentences": "https://www.youtube.com/watch?v=RNUZBHlRH4Y",
    "Question Words (Who/What/Where/When/Why)": "https://www.youtube.com/watch?v=vXWK1-L41f0",
    "3-D Shapes & Symmetry": "https://www.youtube.com/watch?v=2cg-Uc556-Q",
    "The Alphabet A–Z": "https://www.youtube.com/watch?v=tKsIi1MH4lw",
    "Sight Words & Phonics Sounds": "https://www.youtube.com/watch?v=kWtMmRZDY-4",
    "Memory Match & Concentration": "https://www.youtube.com/watch?v=llObUyCEvA0",
    "Sorting, Classifying & Patterns": "https://www.youtube.com/watch?v=MBjjxSx45-Q",
    "Puzzles, Strategy & Creative Drawing": "https://www.youtube.com/watch?v=Norbweaqljc"
}

# New video URLs mapped by the topic names used in curriculum.json (and Excel)
new_links_mapping_old = {
    "Hello, Bye & Basic Colours": "https://www.youtube.com/watch?v=fN1Cyr0ZK9M",
    "Opposites & Simple Words": "https://www.youtube.com/watch?v=RNUZBHlRH4Y",
    "Question Words & Complete Sentences": "https://www.youtube.com/watch?v=vXWK1-L41f0",
    "3-D Shapes & Lines of Symmetry": "https://www.youtube.com/watch?v=2cg-Uc556-Q",
    "Letters A, B, C, D, E": "https://www.youtube.com/watch?v=tKsIi1MH4lw",
    "Phonics Rules, Digraphs & Sight Words": "https://www.youtube.com/watch?v=kWtMmRZDY-4",
    "Match the Pairs (Memory Game)": "https://www.youtube.com/watch?v=llObUyCEvA0",
    "Sorting, Patterns & Simon Says": "https://www.youtube.com/watch?v=MBjjxSx45-Q",
    "Strategy, Puzzles & Creative Drawing": "https://www.youtube.com/watch?v=Norbweaqljc"
}

# 1. Update the Excel Spreadsheet
if os.path.exists(excel_path):
    print(f"Updating Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    if 'Curriculum Dataset' in wb.sheetnames:
        ws = wb['Curriculum Dataset']
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        try:
            topic_col_idx = headers.index('TOPIC') + 1
            video_col_idx = headers.index('VIDEO URL') + 1
            print(f"Found columns: TOPIC (Col {topic_col_idx}), VIDEO URL (Col {video_col_idx})")
            
            updates_count = 0
            # Iterate through rows starting from 2
            for row in range(2, ws.max_row + 1):
                topic_val = ws.cell(row=row, column=topic_col_idx).value
                if topic_val:
                    # Clean it up to strip any extra whitespace or invisible characters
                    topic_cleaned = str(topic_val).strip()
                    if topic_cleaned in new_links_mapping_old:
                        new_url = new_links_mapping_old[topic_cleaned]
                        ws.cell(row=row, column=video_col_idx, value=new_url)
                        print(f"Excel row {row}: Updated '{topic_cleaned}' video URL to: {new_url}")
                        updates_count += 1
            
            wb.save(excel_path)
            print(f"Successfully saved Excel updates! Total cells updated: {updates_count}")
        except ValueError as e:
            print(f"Error finding headers: {e}")
    else:
        print("Sheet 'Curriculum Dataset' not found in Excel workbook.")
else:
    print(f"Excel path {excel_path} not found.")

# 2. Update curriculum_new.json
if os.path.exists(curriculum_new_path):
    print(f"\nUpdating {curriculum_new_path}...")
    with open(curriculum_new_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    updates_count = 0
    for item in data:
        topic_val = item.get('TOPIC')
        if topic_val in new_links_mapping_new:
            new_url = new_links_mapping_new[topic_val]
            item['VIDEO URL'] = new_url
            updates_count += 1
            
    with open(curriculum_new_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"Successfully updated curriculum_new.json. Total items updated: {updates_count}")
else:
    print(f"File {curriculum_new_path} not found.")

# 3. Update curriculum.json
if os.path.exists(curriculum_path):
    print(f"\nUpdating {curriculum_path}...")
    with open(curriculum_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    updates_count = 0
    for item in data:
        topic_val = item.get('topic')
        if topic_val in new_links_mapping_old:
            new_url = new_links_mapping_old[topic_val]
            item['video_url'] = new_url
            updates_count += 1
            
    with open(curriculum_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"Successfully updated curriculum.json. Total items updated: {updates_count}")
else:
    print(f"File {curriculum_path} not found.")

print("\nAll dataset links updated successfully!")
